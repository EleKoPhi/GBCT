from sqlite3 import *
import os
from defines import *
from openpyxl import *
from difflib import *
from userdatacontainer import *
from pathlib import Path


class DataHandler():
    sql_user = """
    CREATE TABLE USER ( 
    USER_NUMBER INTEGER PRIMARY KEY, 
    first_name VARCHAR(50), 
    last_name VARCHAR(50), 
    department CHAR(50), 
    currentID CHAR(20),
    status CHAR(1),
    price CHAR(1));"""

    sql_deposit = """
    CREATE TABLE DEPOSIT (
    DEPOSIT_NUMBER INTEGER PRIMARY KEY,
    amound CHAR(10),
    user_name CHAR(20),
    old_credit CHAR(20),
    new_credit CHAR(20),
    consumption CHAR(20),
    Type CHAR(1),
    date DATE);"""

    sql_id = """
    CREATE TABLE IDS (
    ID_NUMBER INTEGER PRIMARY KEY,
    ID CHAR(20),
    User CHAR(50),
    change_date DATE);"""

    format_str_user = """INSERT INTO USER (USER_NUMBER, first_name, last_name, department, currentID, status, price)
        VALUES (NULL, "{first_name}", "{last_name}", "{department}", "{currentID}", "{status}", "{price}");"""

    format_str_deposit = """INSERT INTO DEPOSIT (DEPOSIT_NUMBER, amound, user_name, old_credit, new_credit, consumption, Type, date)
        VALUES (NULL, "{amound}", "{user_name}", "{old_credit}", "{new_credit}", "{consumption}", "{Type}", "{date}");"""

    format_str_id = """INSERT INTO IDS (ID_NUMBER, ID,User, change_date)
        VALUES (NULL, "{ID}", "{User}", "{change_date}");"""

    def __init__(self):

        self.user_from_SOP = []
        self.user_from_deposit_workbook = []
        self.ids_from_user_excel = []
        self.name_matches = []
        self.users = {}
        self.deposits_raw = None

        base_path = Path(__file__).parent
        db_name = "ES-2_DataBase.db"
        data_base = (base_path / "../Data" / db_name).resolve()
        self.deposits_excel = (base_path / "../Data/Einzahlungen_database.xlsx").resolve()
        self.user_excel = (base_path / "../Data/User.xlsx").resolve()

        remove_flag = 2  # int(input("New DB? (1: YES) (2: NO)"))

        if remove_flag is 1:
            # os.remove(data_base)
            self.read_user_from_excel()
            self.data_base = connect(data_base)
            self.data_base_cursor = self.data_base.cursor()
            self.data_base_cursor.execute(self.sql_user)
            self.data_base_cursor.execute(self.sql_deposit)
            self.data_base_cursor.execute(self.sql_id)
            self.add_dummy_data()

        else:
            self.data_base = connect(data_base)
            self.data_base_cursor = self.data_base.cursor()

        self.update_users_from_db()
        self.update_deposits_from_db()

        self.hand_work()


    def update_deposits_from_db(self):
        self.deposits_raw = self.get_from_db('DEPOSIT')

    def update_users_from_db(self):

        self.users = {}

        for user_raw in self.get_from_db('User ORDER BY first_name ASC'):
            _user = UserDataContainer(user_raw)
            self.users[_user.surname + ' ' + _user.first_name] = _user

    def add_user_to_database(self, first_name=None, last_name=None, department=None, currentID=None, status=None,
                             price=None):

        sql_command = self.format_str_user.format(first_name=first_name, last_name=last_name, department=department,
                                                  currentID=currentID,
                                                  status=status, price=price)

        self.data_base_cursor.execute(sql_command)
        self.data_base.commit()

    def add_deposit_to_database(self, amount=None, user_name=None, old_credit=None, new_credit=None, consumption=None,
                                Type=None, date=None):

        sql_command = self.format_str_deposit.format(amound=amount, user_name=user_name, old_credit=old_credit,
                                                     new_credit=new_credit, consumption=consumption, Type=Type,
                                                     date=date)

        self.data_base_cursor.execute(sql_command)
        self.data_base.commit()

    def add_id_to_database(self, card_id=None, card_user=None, change_date=None):

        sql_command = self.format_str_id.format(ID=card_id, User=card_user, change_date=change_date)

        self.data_base_cursor.execute(sql_command)
        self.data_base.commit()

    def get_from_db(self, Table):

        try:
            sqlite_select_query = """SELECT * FROM {}""".format(Table)
            self.data_base_cursor.execute(sqlite_select_query)
            return self.data_base_cursor.fetchall()

        except:
            print("SQL ERROR")
            return None

    def set_value(self, value, user, key):

        _user_stream = user.split(' ')
        _last_name = _user_stream[0]
        _fist_name = _user_stream[1]

        sqlite_select_query = """UPDATE USER SET %s = '%s' WHERE first_name = '%s' AND last_name = '%s'""" % (
            key, value, _last_name, _fist_name)
        try:

            print(sqlite_select_query)
            self.data_base_cursor.execute(sqlite_select_query)
            self.data_base.commit()

        except:
            print("SQL ERROR")
            return None

    def delete_user(self, user_name):

        user_name = user_name.replace("\t", " ").split(" ")
        first_name = user_name[0]
        last_name =  user_name[1]

        sqlite_select_query = """DELETE FROM USER WHERE first_name = '%s' AND last_name = '%s'""" % (
            first_name, last_name)

        print(sqlite_select_query)

        try:
            self.data_base_cursor.execute(sqlite_select_query)
            self.data_base.commit()

        except:
            print("SQL ERROR")
            return None

    def delete_all_deposits(self):

        sqlite_select_query = """DELETE FROM DEPOSIT"""

        print(sqlite_select_query)

        try:
            self.data_base_cursor.execute(sqlite_select_query)
            self.data_base.commit()

        except:
            print("SQL ERROR")
            return None

    def delete_id(self, _id):

        _id = _id.replace("-", "")
        sqlite_select_query = """DELETE FROM IDS WHERE ID = '%s'""" % _id

        try:
            self.data_base_cursor.execute(sqlite_select_query)
            self.data_base.commit()

        except:
            print("SQL ERROR")
            return None

    def add_dummy_data(self):

        for name in self.name_matches:

            if len(name) == 3:
                first_name = name[0].split(' ')[0]
                second_name = name[0].split(' ')[1]
                personal_id = name[2]

                if name[2] is None:
                    continue

                self.add_user_to_database(first_name, second_name, "unknown", personal_id, active_user_flag,
                                          normal_price_flag)

                print("## " + first_name + " " + second_name + " added to db ##")

                deposits = self.deposits_excel[name[0]]

                row = deposits['A']

                bar = deposits['C']
                pay_pal_column = deposits['D']
                material = deposits['E']

                value = 0
                user_name = name[0]
                type_flag = None

                for iterator, cell in enumerate(row[1:]):

                    if cell.value is None:
                        # print("No deposits from " + user_name)
                        break

                    else:
                        date = str(cell.value)

                        if " " in date:
                            date = date.split(' ')[0]
                            # print("deposit from " + user_name + " at " + date)

                        if str(bar[iterator + 1].value).isdigit():
                            type_flag = Bar
                            value = str(bar[iterator + 1].value)
                            # print("Cash deposit: " + value)

                        elif str(pay_pal_column[iterator + 1].value).isdigit():
                            type_flag = PayPal
                            value = str(pay_pal_column[iterator + 1].value)
                            # print("PayPal deposit: " + value)

                        elif str(material[iterator + 1].value).isdigit():
                            type_flag = Material
                            value = str(material[iterator + 1].value)
                            # print("Material deposit: " + value)
                        else:
                            print(user_name)
                            print(date)
                            type_flag = input()

                        print(type_flag)

                        self.add_deposit_to_database(value, user_name, "0", "0", "0", type_flag, date)

            if len(name) == 2:

                if name[0] == 'Bartel':
                    self.add_user_to_database('Batel', 'Sebastian', "", '147773090', active_user_flag,
                                              normal_price_flag)
                    continue

    def add_users(self):

        self.read_user_from_excel()
        # print(self.user_from_deposit_workbook)

        for user in self.user_from_deposit_workbook:
            first = user.split(" ")[0]
            last = user.split(" ")[1:]

            ln = ""

            if len(last) > 1:

                for name in last:
                    ln += str(name)
                    ln += " "

            else:
                try:
                    ln = last[0]
                except:
                    ln = ""

            cmd = "USER WHERE first_name = '%s' AND last_name = '%s'" % (first, ln)
            data = self.get_from_db(cmd)
            print(data)

            if len(data) == 0:
                self.add_user_to_database(first, ln, "unknown", None, active_user_flag,
                                          normal_price_flag)

    def read_user_from_excel(self):

        self.deposits_excel = load_workbook(filename=self.deposits_excel)
        self.user_excel = load_workbook(filename=self.user_excel)
        self.user_excel_sheet = self.user_excel.active

        _row = self.user_excel_sheet['A']

        for elements_from_row_a in _row[1:]:

            if elements_from_row_a.value is not None:
                self.user_from_SOP.append(elements_from_row_a.value)

        for name_from_deposits_workbook in self.deposits_excel.sheetnames:

            if name_from_deposits_workbook is not None:
                self.user_from_deposit_workbook.append(name_from_deposits_workbook)

        _row = self.user_excel_sheet['D']

        for id_from_row_d in _row[1:]:

            if id_from_row_d.value is None:
                self.ids_from_user_excel.append(id_from_row_d.value)
                continue

            id_from_row_d.value = str(id_from_row_d.value)

            if '-' in id_from_row_d.value:
                self.ids_from_user_excel.append(id_from_row_d.value.replace('-', ''))
            else:
                self.ids_from_user_excel.append(id_from_row_d.value)

        for user in self.user_from_SOP:
            temp = get_close_matches(user, self.user_from_deposit_workbook)
            temp.append(user)

            self.name_matches.append(temp)

        for count, item in enumerate(self.name_matches, start=0):
            self.name_matches[count].append(self.ids_from_user_excel[count])

    def hand_work(self):

        #sqlite_select_query = """UPDATE USER SET %s = '%s' WHERE first_name = '%s' AND last_name = '%s'""" % ("currentID", "None", "Dippel", "Arnd")
        """try:
            print(sqlite_select_query)
            self.data_base_cursor.execute(sqlite_select_query)
            self.data_base.commit()
        except:
            print("SQL ERROR")
            return None"""

        None



#a=DataHandler()

#print(a.delete_all_deposits())



